"""
Dashboard Creation Script for EG Retail Sales Case Study
Creates an Excel dashboard with all required KPIs and visualizations

Requirements from Part B:
1. Total sales revenue (monthly/quarterly) + top 3 orders
2. Average Order Value (AOV) + products above average
3. Average delivery time + delivery delays
4. Top 3 revenue categories + underperforming products
5. Unpaid orders % + revenue at risk
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

print("="*60)
print("CREATING EXCEL DASHBOARD")
print("="*60)

# ============================================
# 1. LOAD BI-READY DATASET
# ============================================
print("\n1. Loading BI-Ready dataset...")

try:
    bi_sales = pd.read_excel("/Users/instabug/Downloads/salma/BI_Ready_Sales_Dataset.xlsx")
    print(f"✓ Loaded {len(bi_sales)} records")
    print(f"✓ Columns: {len(bi_sales.columns)}")
except FileNotFoundError:
    print("❌ Error: BI_Ready_Sales_Dataset.xlsx not found!")
    print("Please run Retail_Sales_Cleaned.py first to generate the BI dataset.")
    exit(1)

# ============================================
# 2. CALCULATE KEY METRICS
# ============================================
print("\n2. Calculating KPIs...")

# 2.1 Total Sales Revenue
total_revenue = bi_sales['TotalAmount_Calc'].sum()
print(f"   Total Revenue: {total_revenue:,.2f} EGP")

# 2.2 Revenue by Month
revenue_by_month = bi_sales.groupby('Order_YearMonth')['TotalAmount_Calc'].sum().reset_index()
revenue_by_month.columns = ['Month', 'Revenue']
revenue_by_month = revenue_by_month.sort_values('Month')
print(f"   Months analyzed: {len(revenue_by_month)}")

# 2.3 Revenue by Quarter
revenue_by_quarter = bi_sales.groupby(['Order_Year', 'Order_Quarter'])['TotalAmount_Calc'].sum().reset_index()
revenue_by_quarter['Quarter'] = revenue_by_quarter['Order_Year'].astype(str) + '-Q' + revenue_by_quarter['Order_Quarter'].astype(str)
revenue_by_quarter = revenue_by_quarter[['Quarter', 'TotalAmount_Calc']]
revenue_by_quarter.columns = ['Quarter', 'Revenue']
print(f"   Quarters analyzed: {len(revenue_by_quarter)}")

# 2.4 Top 3 Orders by Revenue
top_3_orders = bi_sales.nlargest(3, 'TotalAmount_Calc')[['OrderID_cleaned', 'TotalAmount_Calc', 'Order_YearMonth', 'ProductName_Clean', 'CustomerName_clean']]
print(f"   Top 3 orders identified")

# 2.5 Average Order Value (AOV)
total_orders = bi_sales['OrderID_cleaned'].nunique()
aov = total_revenue / total_orders
print(f"   Average Order Value: {aov:,.2f} EGP")

# 2.6 Products with Above-Average AOV
product_aov = bi_sales.groupby('ProductName_Clean')['TotalAmount_Calc'].mean().reset_index()
product_aov.columns = ['Product', 'AOV']
product_aov['Above_Average'] = product_aov['AOV'] > aov
products_above_aov = product_aov[product_aov['Above_Average']].sort_values('AOV', ascending=False)
print(f"   Products above AOV: {len(products_above_aov)}")

# 2.7 Average Delivery Time
avg_delivery_time = bi_sales['Delivery_Time_Days'].mean()
print(f"   Avg Delivery Time: {avg_delivery_time:.1f} days")

# 2.8 Delayed Deliveries (>5 days)
delayed_deliveries = bi_sales[bi_sales['Delivery_Delayed'] == True]
delayed_pct = len(delayed_deliveries) / len(bi_sales) * 100
print(f"   Delayed orders: {len(delayed_deliveries)} ({delayed_pct:.1f}%)")

# 2.9 Top 3 Categories by Revenue
category_revenue = bi_sales.groupby('Category_Clean')['TotalAmount_Calc'].sum().reset_index()
category_revenue.columns = ['Category', 'Revenue']
top_3_categories = category_revenue.nlargest(3, 'Revenue')
print(f"   Top 3 categories identified")

# 2.10 Product Performance (Quantity sold)
product_performance = bi_sales.groupby('ProductName_Clean').agg({
    'Quantity_Clean': 'sum',
    'TotalAmount_Calc': 'sum'
}).reset_index()
product_performance.columns = ['Product', 'Quantity_Sold', 'Revenue']
product_performance = product_performance.sort_values('Quantity_Sold', ascending=True)
underperforming_products = product_performance.head(10)  # Bottom 10
print(f"   Underperforming products identified")

# 2.11 Unpaid Orders Analysis
unpaid_orders = bi_sales[bi_sales['PaymentStatus_Clean'] == 'Unpaid']
unpaid_count = len(unpaid_orders)
unpaid_pct = unpaid_count / len(bi_sales) * 100
revenue_at_risk = unpaid_orders['TotalAmount_Calc'].sum()
print(f"   Unpaid orders: {unpaid_count} ({unpaid_pct:.1f}%)")
print(f"   Revenue at risk: {revenue_at_risk:,.2f} EGP")

# ============================================
# 3. CREATE EXCEL WORKBOOK
# ============================================
print("\n3. Creating Excel workbook...")

wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
kpi_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
kpi_font = Font(bold=True, size=14)
title_font = Font(bold=True, size=16)
center_align = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================
# SHEET 1: EXECUTIVE SUMMARY
# ============================================
print("   Creating Executive Summary sheet...")
ws_summary = wb.create_sheet("Executive Summary")

# Title
ws_summary['A1'] = "EG RETAIL SALES DASHBOARD"
ws_summary['A1'].font = title_font
ws_summary.merge_cells('A1:D1')

ws_summary['A2'] = "Executive Summary - Key Performance Indicators"
ws_summary['A2'].font = Font(size=12, italic=True)
ws_summary.merge_cells('A2:D2')

# KPI Section
row = 4
ws_summary[f'A{row}'] = "KEY METRICS"
ws_summary[f'A{row}'].font = Font(bold=True, size=14)
ws_summary.merge_cells(f'A{row}:D{row}')
ws_summary[f'A{row}'].fill = header_fill
ws_summary[f'A{row}'].font = header_font

# KPI Values
kpis = [
    ("Total Revenue (EGP)", f"{total_revenue:,.2f}"),
    ("Total Orders", f"{total_orders:,}"),
    ("Average Order Value (EGP)", f"{aov:,.2f}"),
    ("Average Delivery Time (days)", f"{avg_delivery_time:.1f}"),
    ("Delayed Deliveries (%)", f"{delayed_pct:.1f}%"),
    ("Unpaid Orders (%)", f"{unpaid_pct:.1f}%"),
    ("Revenue at Risk (EGP)", f"{revenue_at_risk:,.2f}"),
]

row = 5
for kpi_name, kpi_value in kpis:
    ws_summary[f'A{row}'] = kpi_name
    ws_summary[f'B{row}'] = kpi_value
    ws_summary[f'A{row}'].fill = kpi_fill
    ws_summary[f'B{row}'].font = kpi_font
    ws_summary[f'B{row}'].alignment = Alignment(horizontal="right")
    row += 1

# Top 3 Orders Section
row += 2
ws_summary[f'A{row}'] = "TOP 3 ORDERS BY REVENUE"
ws_summary[f'A{row}'].font = Font(bold=True, size=14)
ws_summary.merge_cells(f'A{row}:E{row}')
ws_summary[f'A{row}'].fill = header_fill
ws_summary[f'A{row}'].font = header_font

row += 1
headers = ['Rank', 'Order ID', 'Revenue (EGP)', 'Month', 'Product']
for col_idx, header in enumerate(headers, start=1):
    cell = ws_summary.cell(row, col_idx, header)
    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    cell.font = Font(bold=True)
    cell.border = border

for idx, (_, order_row) in enumerate(top_3_orders.iterrows(), start=1):
    row += 1
    ws_summary.cell(row, 1, idx)
    ws_summary.cell(row, 2, order_row['OrderID_cleaned'])
    ws_summary.cell(row, 3, f"{order_row['TotalAmount_Calc']:,.2f}")
    ws_summary.cell(row, 4, str(order_row['Order_YearMonth']))
    ws_summary.cell(row, 5, order_row['ProductName_Clean'])

# Top 3 Categories
row += 3
ws_summary[f'A{row}'] = "TOP 3 CATEGORIES BY REVENUE"
ws_summary[f'A{row}'].font = Font(bold=True, size=14)
ws_summary.merge_cells(f'A{row}:C{row}')
ws_summary[f'A{row}'].fill = header_fill
ws_summary[f'A{row}'].font = header_font

row += 1
ws_summary.cell(row, 1, 'Rank').font = Font(bold=True)
ws_summary.cell(row, 2, 'Category').font = Font(bold=True)
ws_summary.cell(row, 3, 'Revenue (EGP)').font = Font(bold=True)

for idx, (_, cat_row) in enumerate(top_3_categories.iterrows(), start=1):
    row += 1
    ws_summary.cell(row, 1, idx)
    ws_summary.cell(row, 2, cat_row['Category'])
    ws_summary.cell(row, 3, f"{cat_row['Revenue']:,.2f}")

# Adjust column widths
ws_summary.column_dimensions['A'].width = 30
ws_summary.column_dimensions['B'].width = 20
ws_summary.column_dimensions['C'].width = 20
ws_summary.column_dimensions['D'].width = 15
ws_summary.column_dimensions['E'].width = 30

# ============================================
# SHEET 2: REVENUE ANALYSIS
# ============================================
print("   Creating Revenue Analysis sheet...")
ws_revenue = wb.create_sheet("Revenue Analysis")

# Title
ws_revenue['A1'] = "REVENUE ANALYSIS"
ws_revenue['A1'].font = title_font
ws_revenue.merge_cells('A1:C1')

# Monthly Revenue
ws_revenue['A3'] = "Monthly Revenue"
ws_revenue['A3'].font = Font(bold=True, size=12)

ws_revenue.cell(4, 1, 'Month').font = Font(bold=True)
ws_revenue.cell(4, 2, 'Revenue (EGP)').font = Font(bold=True)

for idx, (_, month_row) in enumerate(revenue_by_month.iterrows(), start=5):
    ws_revenue.cell(idx, 1, month_row['Month'])
    ws_revenue.cell(idx, 2, month_row['Revenue'])

# Quarterly Revenue
col_start = 4
ws_revenue.cell(3, col_start, "Quarterly Revenue").font = Font(bold=True, size=12)
ws_revenue.cell(4, col_start, 'Quarter').font = Font(bold=True)
ws_revenue.cell(4, col_start+1, 'Revenue (EGP)').font = Font(bold=True)

for idx, (_, quarter_row) in enumerate(revenue_by_quarter.iterrows(), start=5):
    ws_revenue.cell(idx, col_start, quarter_row['Quarter'])
    ws_revenue.cell(idx, col_start+1, quarter_row['Revenue'])

# Category Revenue
ws_revenue.cell(len(revenue_by_month) + 7, 1, "Revenue by Category").font = Font(bold=True, size=12)
cat_start_row = len(revenue_by_month) + 8

ws_revenue.cell(cat_start_row, 1, 'Category').font = Font(bold=True)
ws_revenue.cell(cat_start_row, 2, 'Revenue (EGP)').font = Font(bold=True)

for idx, (_, cat_row) in enumerate(category_revenue.sort_values('Revenue', ascending=False).iterrows(), start=cat_start_row+1):
    ws_revenue.cell(idx, 1, cat_row['Category'])
    ws_revenue.cell(idx, 2, cat_row['Revenue'])

ws_revenue.column_dimensions['A'].width = 20
ws_revenue.column_dimensions['B'].width = 20
ws_revenue.column_dimensions['D'].width = 20
ws_revenue.column_dimensions['E'].width = 20

# ============================================
# SHEET 3: PRODUCT PERFORMANCE
# ============================================
print("   Creating Product Performance sheet...")
ws_products = wb.create_sheet("Product Performance")

ws_products['A1'] = "PRODUCT PERFORMANCE ANALYSIS"
ws_products['A1'].font = title_font
ws_products.merge_cells('A1:D1')

# Products Above AOV
ws_products['A3'] = f"Products with Above-Average Order Value (AOV > {aov:,.2f} EGP)"
ws_products['A3'].font = Font(bold=True, size=12)
ws_products.merge_cells('A3:C3')

ws_products.cell(4, 1, 'Product').font = Font(bold=True)
ws_products.cell(4, 2, 'Average Order Value (EGP)').font = Font(bold=True)
ws_products.cell(4, 3, 'Above Average?').font = Font(bold=True)

for idx, (_, prod_row) in enumerate(products_above_aov.iterrows(), start=5):
    ws_products.cell(idx, 1, prod_row['Product'])
    ws_products.cell(idx, 2, prod_row['AOV'])
    ws_products.cell(idx, 3, 'Yes')

# Underperforming Products
under_start_row = len(products_above_aov) + 7
ws_products.cell(under_start_row, 1, "Underperforming Products (Lowest 10 by Quantity Sold)").font = Font(bold=True, size=12)
ws_products.merge_cells(f'A{under_start_row}:D{under_start_row}')

ws_products.cell(under_start_row+1, 1, 'Product').font = Font(bold=True)
ws_products.cell(under_start_row+1, 2, 'Quantity Sold').font = Font(bold=True)
ws_products.cell(under_start_row+1, 3, 'Revenue (EGP)').font = Font(bold=True)

for idx, (_, prod_row) in enumerate(underperforming_products.iterrows(), start=under_start_row+2):
    ws_products.cell(idx, 1, prod_row['Product'])
    ws_products.cell(idx, 2, prod_row['Quantity_Sold'])
    ws_products.cell(idx, 3, prod_row['Revenue'])

ws_products.column_dimensions['A'].width = 35
ws_products.column_dimensions['B'].width = 25
ws_products.column_dimensions['C'].width = 20

# ============================================
# SHEET 4: DELIVERY PERFORMANCE
# ============================================
print("   Creating Delivery Performance sheet...")
ws_delivery = wb.create_sheet("Delivery Performance")

ws_delivery['A1'] = "DELIVERY PERFORMANCE ANALYSIS"
ws_delivery['A1'].font = title_font
ws_delivery.merge_cells('A1:D1')

# Summary Metrics
ws_delivery['A3'] = "Delivery Metrics"
ws_delivery['A3'].font = Font(bold=True, size=12)

ws_delivery.cell(4, 1, 'Metric').font = Font(bold=True)
ws_delivery.cell(4, 2, 'Value').font = Font(bold=True)

delivery_metrics = [
    ("Average Delivery Time (days)", f"{avg_delivery_time:.1f}"),
    ("Delayed Deliveries (>5 days)", f"{len(delayed_deliveries)}"),
    ("Delayed Delivery Rate (%)", f"{delayed_pct:.1f}%"),
    ("On-Time Deliveries", f"{len(bi_sales) - len(delayed_deliveries)}"),
    ("On-Time Rate (%)", f"{100 - delayed_pct:.1f}%"),
]

for idx, (metric, value) in enumerate(delivery_metrics, start=5):
    ws_delivery.cell(idx, 1, metric)
    ws_delivery.cell(idx, 2, value)

# Delayed Orders List
ws_delivery.cell(12, 1, "Delayed Orders (Delivery Time > 5 days)").font = Font(bold=True, size=12)
ws_delivery.merge_cells('A12:E12')

ws_delivery.cell(13, 1, 'Order ID').font = Font(bold=True)
ws_delivery.cell(13, 2, 'Delivery Time (days)').font = Font(bold=True)
ws_delivery.cell(13, 3, 'Product').font = Font(bold=True)
ws_delivery.cell(13, 4, 'Governorate').font = Font(bold=True)

for idx, (_, del_row) in enumerate(delayed_deliveries.head(20).iterrows(), start=14):
    ws_delivery.cell(idx, 1, del_row['OrderID_cleaned'])
    ws_delivery.cell(idx, 2, del_row['Delivery_Time_Days'])
    ws_delivery.cell(idx, 3, del_row['ProductName_Clean'])
    ws_delivery.cell(idx, 4, del_row['Governorate_Clean'])

ws_delivery.column_dimensions['A'].width = 20
ws_delivery.column_dimensions['B'].width = 25
ws_delivery.column_dimensions['C'].width = 35
ws_delivery.column_dimensions['D'].width = 20

# ============================================
# SHEET 5: PAYMENT & RISK ANALYSIS
# ============================================
print("   Creating Payment & Risk Analysis sheet...")
ws_payment = wb.create_sheet("Payment & Risk Analysis")

ws_payment['A1'] = "PAYMENT STATUS & RISK ANALYSIS"
ws_payment['A1'].font = title_font
ws_payment.merge_cells('A1:D1')

# Payment Status Summary
payment_summary = bi_sales.groupby('PaymentStatus_Clean').agg({
    'OrderID_cleaned': 'count',
    'TotalAmount_Calc': 'sum'
}).reset_index()
payment_summary.columns = ['Payment Status', 'Order Count', 'Total Amount (EGP)']
payment_summary['Percentage (%)'] = payment_summary['Order Count'] / len(bi_sales) * 100

ws_payment['A3'] = "Payment Status Breakdown"
ws_payment['A3'].font = Font(bold=True, size=12)

ws_payment.cell(4, 1, 'Payment Status').font = Font(bold=True)
ws_payment.cell(4, 2, 'Order Count').font = Font(bold=True)
ws_payment.cell(4, 3, 'Total Amount (EGP)').font = Font(bold=True)
ws_payment.cell(4, 4, 'Percentage (%)').font = Font(bold=True)

for idx, (_, pay_row) in enumerate(payment_summary.iterrows(), start=5):
    ws_payment.cell(idx, 1, pay_row['Payment Status'])
    ws_payment.cell(idx, 2, pay_row['Order Count'])
    ws_payment.cell(idx, 3, f"{pay_row['Total Amount (EGP)']:,.2f}")
    ws_payment.cell(idx, 4, f"{pay_row['Percentage (%)']:.1f}%")

# Risk Analysis
row = len(payment_summary) + 7
ws_payment.cell(row, 1, "RISK ANALYSIS").font = Font(bold=True, size=12)
ws_payment.cell(row, 1).fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
ws_payment.merge_cells(f'A{row}:B{row}')

risk_metrics = [
    ("Unpaid Orders Count", f"{unpaid_count}"),
    ("Unpaid Orders (%)", f"{unpaid_pct:.1f}%"),
    ("Revenue at Risk (EGP)", f"{revenue_at_risk:,.2f}"),
    ("Percentage of Total Revenue at Risk (%)", f"{revenue_at_risk/total_revenue*100:.1f}%"),
]

row += 1
for metric, value in risk_metrics:
    ws_payment.cell(row, 1, metric)
    ws_payment.cell(row, 2, value)
    ws_payment.cell(row, 2).font = Font(bold=True, color="FF0000")
    row += 1

# Unpaid Orders List
row += 2
ws_payment.cell(row, 1, "Unpaid Orders (Top 20 by Amount)").font = Font(bold=True, size=12)
ws_payment.merge_cells(f'A{row}:E{row}')

row += 1
ws_payment.cell(row, 1, 'Order ID').font = Font(bold=True)
ws_payment.cell(row, 2, 'Amount (EGP)').font = Font(bold=True)
ws_payment.cell(row, 3, 'Customer').font = Font(bold=True)
ws_payment.cell(row, 4, 'Status').font = Font(bold=True)
ws_payment.cell(row, 5, 'Days Since Order').font = Font(bold=True)

unpaid_top = unpaid_orders.nlargest(20, 'TotalAmount_Calc')
for _, unpaid_row in unpaid_top.iterrows():
    row += 1
    ws_payment.cell(row, 1, unpaid_row['OrderID_cleaned'])
    ws_payment.cell(row, 2, f"{unpaid_row['TotalAmount_Calc']:,.2f}")
    ws_payment.cell(row, 3, unpaid_row['CustomerName_clean'])
    ws_payment.cell(row, 4, unpaid_row['Status_Clean'])
    # Calculate days since order (if we have the date)
    if pd.notna(unpaid_row['OrderDate']):
        days_since = (pd.Timestamp.now() - unpaid_row['OrderDate']).days
        ws_payment.cell(row, 5, days_since)

ws_payment.column_dimensions['A'].width = 25
ws_payment.column_dimensions['B'].width = 20
ws_payment.column_dimensions['C'].width = 25
ws_payment.column_dimensions['D'].width = 20
ws_payment.column_dimensions['E'].width = 20

# ============================================
# 4. SAVE WORKBOOK
# ============================================
print("\n4. Saving dashboard...")

output_file = "/Users/instabug/Downloads/salma/EG_Retail_Sales_Dashboard.xlsx"
wb.save(output_file)

print(f"\n✅ Dashboard created successfully!")
print(f"📊 File saved to: {output_file}")
print("\n" + "="*60)
print("DASHBOARD CONTENTS:")
print("="*60)
print("Sheet 1: Executive Summary - KPIs and top performers")
print("Sheet 2: Revenue Analysis - Monthly, quarterly, category breakdown")
print("Sheet 3: Product Performance - Above-AOV products, underperformers")
print("Sheet 4: Delivery Performance - Avg time, delays, SLA tracking")
print("Sheet 5: Payment & Risk Analysis - Unpaid orders, revenue at risk")
print("="*60)
print("\n✅ Ready for submission!\n")

